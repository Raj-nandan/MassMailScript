
import smtplib
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Load emails from Excel (assumes file has a column named 'Email')
wb = load_workbook("hr_contacts.xlsx")
sheet = wb.active

emails = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header row
    if row[0]:  # assuming 'Email' is the first column
        emails.append(row[0])

# Gmail credentials
YOUR_EMAIL = "enter your email"
YOUR_PASSWORD = "xxxx xxxx xxxx xxxx"  # Use Gmail App Password

# Email content
subject = "Application for Software Developer – 2025 Graduate"
body = """\
Dear Hiring Manager,

I hope this message finds you well.

My name is Raj Nandan, 2025 B.tech Graduate in Information Technology. I am reaching out to express my keen interest in entry-level opportunities in Software Development within your esteemed organization.

During my academic journey and internship experiences, I have built a strong foundation in full-stack development, system design, and DevOps practices. Most recently, as an SRE Intern at BARQ, I worked closely with DevOps and Kubernetes teams to monitor and optimize live production systems using Grafana and Elastic, ensuring reliability and high availability.

Beyond my internship, I have led the development of scalable applications such as:
- TimeSync Pro – a MERN-based real-time attendance and monitoring platform.
- Flora Fusion – a collaborative knowledge-sharing system built on Node.js and MongoDB.

My technical expertise includes:
- Languages & Frameworks: JavaScript, TypeScript, Node.js, React.js, C++, Java, Python
- Databases: MongoDB, SQL, NoSQL
- DevOps & Tools: Docker, CI/CD pipelines, Git, REST APIs
- Problem Solving: 500+ coding challenges solved across LeetCode and GeeksforGeeks

I am confident that my skills, coupled with my enthusiasm for building impactful solutions, will allow me to contribute effectively to your team. Please find my resume attached for your review.

Thank you very much for your time and consideration. I look forward to the possibility of contributing to your organization.

Best regards,
Raj Nandan
+91 8409452277
LinkedIn: https://www.linkedin.com/in/raj-nandan/
"""

# Setup SMTP
server = smtplib.SMTP("smtp.gmail.com", 587)
server.starttls()
server.login(YOUR_EMAIL, YOUR_PASSWORD)

# Loop through HR emails
for email in emails:
    msg = MIMEMultipart()
    msg["From"] = YOUR_EMAIL
    msg["To"] = email
    msg["Subject"] = subject
    
    msg.attach(MIMEText(body, "plain"))
    
    # Attach resume
    with open("resume.pdf", "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment; filename=resume.pdf")
        msg.attach(part)
    
    try:
        server.sendmail(YOUR_EMAIL, email, msg.as_string())
        print(f"✅ Email sent to {email}")
    except Exception as e:
        print(f"❌ Failed to send to {email}: {e}")

server.quit()




